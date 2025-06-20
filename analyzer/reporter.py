# /analyzer/reporter.py
import pandas as pd
from datetime import datetime
import os
import subprocess

class ExcelReporter:
    """
    분석 결과를 받아 최종 엑셀 파일을 생성하는 역할만 전담합니다.
    (Responsible for creating the final Excel file from the analysis results.)
    """
    def __init__(self, reports_dict):
        self.reports = reports_dict

    def _format_sheet(self, df, writer, sheet_name):
        """Helper function to format a single Excel sheet."""
        if df.empty: return
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        header_format = workbook.add_format({'bold': True, 'text_wrap': True, 'valign': 'top', 'fg_color': '#D7E4BC', 'border': 1, 'align': 'center'})
        
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
        
        # 컬럼 너비 자동 조정
        for i, col in enumerate(df.columns):
            # 컬럼 내용 중 가장 긴 값과 헤더 이름 중 더 긴 것을 기준으로 너비 설정
            column_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
            worksheet.set_column(i, i, column_len)

    def create_report(self):
        """정규화된 데이터프레임들을 엑셀 리포트로 생성합니다."""
        if not self.reports:
            print("⚠️ No reports to generate.")
            return
        
        # 타임스탬프가 포함된 파일명 생성
        timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
        filename = f"outputs/업로드_스타일_공급사별_Case집계_{timestamp}.xlsx"
        
        # outputs 디렉토리가 없으면 생성
        os.makedirs('outputs', exist_ok=True)
        
        print(f"📊 Creating Excel report: {filename}")
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            for sheet_name, df in self.reports.items():
                if df is not None and not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # 워크시트 가져오기 및 스타일링
                    worksheet = writer.sheets[sheet_name]
                    self._apply_sheet_styling(worksheet, df)
                    
                    print(f"   ✅ Sheet '{sheet_name}' created with {len(df)} rows")
                else:
                    print(f"   ⚠️ Sheet '{sheet_name}' skipped (empty data)")
        
        print(f"🎉 Excel report successfully created: {filename}")
        return filename

    def _apply_sheet_formatting(self, sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
        """시트별 특별 포맷팅을 적용합니다."""
        formatted_df = df.copy()
        
        if sheet_name == 'Consolidated_Status':
            # 월별 데이터 정렬
            if 'Month' in formatted_df.columns:
                if pd.api.types.is_period_dtype(formatted_df['Month']):
                    # Period 타입을 문자열로 변환
                    formatted_df['Month'] = formatted_df['Month'].astype(str)
                elif pd.api.types.is_datetime64_any_dtype(formatted_df['Month']):
                    # Datetime 타입을 월별 문자열로 변환
                    formatted_df['Month'] = formatted_df['Month'].dt.strftime('%Y-%m')
                
                formatted_df = formatted_df.sort_values(['Month', 'Supplier'])
            
            # 숫자 컬럼 포맷팅
            numeric_cols = formatted_df.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                formatted_df[col] = formatted_df[col].round(2)
        
        elif sheet_name == 'Overall_Supplier_Summary':
            # 공급사별 집계 정렬
            if 'Supplier' in formatted_df.columns:
                formatted_df = formatted_df.sort_values('Supplier')
            
            # 숫자 컬럼 포맷팅
            numeric_cols = formatted_df.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                formatted_df[col] = formatted_df[col].round(2)
        
        elif sheet_name == 'Warehouse_Stock_Summary':
            # 창고별 정렬
            if 'Warehouse' in formatted_df.columns:
                formatted_df = formatted_df.sort_values(['Supplier', 'Warehouse'])
            
            # 숫자 컬럼 포맷팅
            numeric_cols = formatted_df.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                formatted_df[col] = formatted_df[col].round(2)
        
        elif sheet_name == 'Pivoted_Monthly_Summary':
            # 월별·분류별 정렬
            if 'Month' in formatted_df.columns:
                if pd.api.types.is_period_dtype(formatted_df['Month']):
                    formatted_df['Month'] = formatted_df['Month'].astype(str)
                elif pd.api.types.is_datetime64_any_dtype(formatted_df['Month']):
                    formatted_df['Month'] = formatted_df['Month'].dt.strftime('%Y-%m')
                    
                formatted_df = formatted_df.sort_values(['Month', 'Classification', 'Supplier'])
            
            # 숫자 컬럼 포맷팅
            numeric_cols = formatted_df.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                formatted_df[col] = formatted_df[col].round(2)
        
        elif sheet_name == 'DeadStock_Analysis (90+ days)':
            # 장기재고 정렬 (일수 내림차순)
            if 'Days_Passed' in formatted_df.columns:
                formatted_df = formatted_df.sort_values('Days_Passed', ascending=False)
            
            # 날짜 포맷팅
            if 'Last_Arrival_Date' in formatted_df.columns:
                formatted_df['Last_Arrival_Date'] = pd.to_datetime(formatted_df['Last_Arrival_Date']).dt.strftime('%Y-%m-%d')
            
            # 숫자 컬럼 포맷팅
            numeric_cols = formatted_df.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                if col != 'Days_Passed':  # 일수는 정수로 유지
                    formatted_df[col] = formatted_df[col].round(2)
        
        elif sheet_name == 'HVDC_Stock_OnHand':
            # 케이스 번호별 정렬
            if 'case_no' in formatted_df.columns:
                formatted_df = formatted_df.sort_values('case_no')
            
            # 숫자 컬럼 포맷팅
            numeric_cols = formatted_df.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                formatted_df[col] = formatted_df[col].round(2)
        
        # === 스키마 v1.2 추가 시트 포맷팅 ===
        elif sheet_name == 'HS_Code_Analysis':
            # HS 코드별 정렬
            if 'hs_code' in formatted_df.columns:
                formatted_df = formatted_df.sort_values(['hs_code', 'Supplier'])
            
            # 숫자 컬럼 포맷팅
            numeric_cols = formatted_df.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                if col != 'case_no':  # case_no는 정수로 유지
                    formatted_df[col] = formatted_df[col].round(2)
        
        elif sheet_name == 'Incoterm_Analysis':
            # Incoterm별 정렬
            if 'incoterm' in formatted_df.columns:
                formatted_df = formatted_df.sort_values(['incoterm', 'Supplier'])
            
            # 숫자 컬럼 포맷팅
            numeric_cols = formatted_df.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                if col != 'case_no':
                    formatted_df[col] = formatted_df[col].round(2)
        
        elif sheet_name == 'OOG_Analysis':
            # OOG 플래그별 정렬
            if 'oog_flag' in formatted_df.columns:
                formatted_df = formatted_df.sort_values(['oog_flag', 'Supplier'])
            
            # 숫자 컬럼 포맷팅
            numeric_cols = formatted_df.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                if col != 'case_no':
                    formatted_df[col] = formatted_df[col].round(2)
        
        elif sheet_name == 'Package_Type_Analysis':
            # 패키지 타입별 정렬
            if 'package_type' in formatted_df.columns:
                formatted_df = formatted_df.sort_values(['package_type', 'Supplier'])
            
            # 숫자 컬럼 포맷팅
            numeric_cols = formatted_df.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                if col != 'case_no':
                    formatted_df[col] = formatted_df[col].round(2)
        
        elif sheet_name == 'Storage_Type_Analysis':
            # 저장 타입별 정렬
            if 'storage_type' in formatted_df.columns:
                formatted_df = formatted_df.sort_values(['storage_type', 'Supplier'])
            
            # 숫자 컬럼 포맷팅
            numeric_cols = formatted_df.select_dtypes(include=['number']).columns
            for col in numeric_cols:
                if col != 'case_no':
                    formatted_df[col] = formatted_df[col].round(2)
        
        return formatted_df

    def _apply_sheet_styling(self, worksheet, df: pd.DataFrame):
        """시트별 스타일링을 적용합니다."""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            worksheet.column_dimensions[get_column_letter(col_idx)].width = len(col_name) + 4
        
        # TOTAL 행 스타일
        total_row_font = Font(bold=True)
        if not df.empty and '월' in df.columns and df['월'].iloc[-1] == 'TOTAL':
            for col_idx in range(1, len(df.columns) + 1):
                worksheet.cell(row=len(df) + 1, column=col_idx).font = total_row_font

    def get_report_summary(self):
        """생성된 각 시트의 행 개수를 요약해서 반환합니다."""
        return {sheet_name: len(df) if hasattr(df, '__len__') else 0 for sheet_name, df in self.reports.items()}
